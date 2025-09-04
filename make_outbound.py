#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_outbound.py  (2025-09-03 v6 – SN 不足逻辑全面修正)
------------------------------------------------------
修正点 (仅影响「SN 不足行应标蓝 + 文件尾提示」)：

* **是否“需要 SN”** 改为在初始化阶段记录：
  `codes_with_sn`, `models_with_sn` —— 只要库存里曾经有过 SN 记录，就视为“需要 SN”，不管后面被分配光。
* 循环中：
  1. 判断 `need_sn = code in codes_with_sn or model in models_with_sn`。
  2. 若 `need_sn`：调用 `allocate_sns`，无论还能不能分配到 SN，都要输出一行；若不足 → 行标蓝，文件尾追加 “SN码不足”。
  3. 若 **不 need_sn**：直接输出行，SN 列空白，不标色。

这样即使同一编码前面把 SN 用光，后续仍会正确判定为 “SN不足” 而标蓝。
（其余逻辑不变：编码/型号皆不存在 → 红行 + 提示）
"""

import sys
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# ---------- 配置 ----------
STORE_LIST = [
    "販売一丁目 Qoo10店",
    "販売一丁目 Amazon店",
    "販売一丁目 Yahoo！ショッピング店",
    "販売一丁目（楽天）",
    "ニューライフ",
    "販売一丁目 Wowma店",
]
HEADER = ["存货编码", "仓库", "数量", "单价", "SN码", "备注"]
ERROR_FILL = PatternFill("solid", fgColor="CCE5FF")  # 蓝
SHORT_FILL = PatternFill("solid", fgColor="FFC7CE")  # 红
ALIAS = {
    "存货编码": "code", "商品编码": "code", "在庫品番": "code",
    "规格型号": "model", "型番": "model",
    "JANコード": "jan",
    "SN码": "sn", "シリアル番号": "sn", "SN": "sn",
}
# --------------------------

def norm(c):
    return ALIAS.get(c.strip(), c.strip())

# ---------- 读取 ----------

def read_orders(path):
    for enc in ("utf-8-sig", "cp932", "utf-16", "windows-1252"):
        try:
            df = pd.read_csv(path, dtype=str, encoding=enc, keep_default_na=False)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SystemExit("无法识别当日発送.csv 的编码，请保存为 UTF-8。")

    df.columns = [c.strip() for c in df.columns]
    for col in ("数量", "単価"):
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(r"[ \u3000]", "", regex=True).replace("", "0")
    df["数量"] = pd.to_numeric(df["数量"], errors="coerce").fillna(0).astype(int)
    df["単価"] = pd.to_numeric(df["単価"], errors="coerce").fillna(0)
    return df


def read_inventory(path):
    eng = "xlrd" if path.lower().endswith(".xls") else "openpyxl"
    df = pd.read_excel(path, dtype=str, engine=eng, keep_default_na=False)
    df = df.rename(columns={c: norm(c) for c in df.columns})
    return df

# ---------- SN 池 ----------

def build_sn_pool(inv):
    sn_by_code, sn_by_model = defaultdict(list), defaultdict(list)
    model2code = {}
    codes_with_sn, models_with_sn = set(), set()

    for _, r in inv.iterrows():
        code  = str(r.get("code", "")).strip()
        model = str(r.get("model", "")).strip()
        sn    = str(r.get("sn", "")).strip()
        if sn:
            if code:
                sn_by_code[code].append(sn)
                codes_with_sn.add(code)
            if model:
                sn_by_model[model].append(sn)
                models_with_sn.add(model)
        if code and model and model not in model2code:
            model2code[model] = code

    return sn_by_code, sn_by_model, model2code, codes_with_sn, models_with_sn


def allocate_sns(code, model, qty, sn_by_code, sn_by_model):
    pool = sn_by_code.get(code, []).copy()
    if len(pool) < qty and model:
        pool += sn_by_model.get(model, [])
    enough = len(pool) >= qty
    chosen = pool[:qty] if enough else pool
    for sn in chosen:
        if sn in sn_by_code.get(code, []):
            sn_by_code[code].remove(sn)
        if model and sn in sn_by_model.get(model, []):
            sn_by_model[model].remove(sn)
    return ",".join(chosen), enough

# ---------- 生成输出 ----------

def build_output(orders, inv, wh):
    sn_by_code, sn_by_model, model2code, codes_sn, models_sn = build_sn_pool(inv)
    codes_all  = set(inv.get("code",  pd.Series(dtype=str)).astype(str).str.strip())
    models_all = set(inv.get("model", pd.Series(dtype=str)).astype(str).str.strip())

    outputs = {}
    flags = defaultdict(lambda: {"err": False, "short": False})

    for store in STORE_LIST:
        df = orders[orders["店舗名"] == store].copy()
        if df.empty:
            continue
        df = df.sort_values("JANコード", kind="stable")
        rows = []
        for _, r in df.iterrows():
            jan   = str(r["JANコード"]).strip()
            model = str(r.get("规格型号", "")).strip()
            qty   = int(r["数量"])
            price = r["単価"]
            note  = str(r["注文番号"]).strip()

            code = jan
            exists_code  = code in codes_all
            exists_model = model and model in models_all

            if not exists_code and exists_model:
                code = model2code.get(model, code)
                exists_code = code in codes_all

            if not exists_code and not exists_model and jan in models_all:
                code = model2code.get(jan, code)
                model = jan
                exists_code = code in codes_all
                exists_model = True

            if not (exists_code or exists_model):  # 红色条件
                rows.append([code, wh, qty, price, "", note, True, False])
                flags[store]["err"] = True
                continue

            need_sn = (code in codes_sn) or (model in models_sn)
            if not need_sn:  # 商品本就无 SN 需求
                rows.append([code, wh, qty, price, "", note, False, False])
                continue

            sn_str, enough = allocate_sns(code, model, qty, sn_by_code, sn_by_model)
            short = not enough
            if short:
                flags[store]["short"] = True
            rows.append([code, wh, qty, price, sn_str, note, False, short])
        outputs[store] = rows
    return outputs, flags

# ---------- 写 Excel ----------

def write_xlsx(store, rows, flag):
    fname = f"{store}+{len(rows)+1}.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "出库"
    ws.append(HEADER)
    for c in ws[1]: c.font = Font(bold=True)

    for row in rows:
        err = row.pop(); short = row.pop()
        ws.append(row)
        if err:
            for c in ws[ws.max_row]: c.fill = ERROR_FILL
        elif short:
            for c in ws[ws.max_row]: c.fill = SHORT_FILL

    if flag["err"]:   ws.append([]); ws.append(["当前表格中有未找到的JANコード"])
    if flag["short"]: ws.append(["SN码不足"])

    wb.save(fname)
    print("Wrote:", fname)

# ---------- main ----------

def main():
    if len(sys.argv) < 4:
        print("用法: python3 make_outbound.py 当日発送.csv 库存余额表.xls [通販倉庫|なんば倉庫]")
        sys.exit(1)
    orders_csv, inv_xls, wh = sys.argv[1:4]
    if wh not in ("通販倉庫", "なんば倉庫"):
        raise SystemExit("仓库参数必须是『通販倉庫』或『なんば倉庫』。")

    orders = read_orders(orders_csv)
    inv    = read_inventory(inv_xls)

    outs, flags = build_output(orders, inv, wh)
    for st in STORE_LIST:
        if st not in outs or len(outs[st]) == 0:
            print(f"\033[31m今天此店铺没有订单: {st}\033[0m"); continue
        write_xlsx(st, outs[st], flags[st])

if __name__ == "__main__":
    main()
