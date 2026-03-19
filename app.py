import streamlit as st
import tempfile
import os
import sys
import subprocess
from pathlib import Path
import io
import zipfile
from openpyxl import load_workbook

st.set_page_config(page_title="梦想云选码工具", page_icon="📦")
st.title("梦想云选码工具")

st.markdown(
    """**使用前注意事项**

1. **库存余额表**：请先删除表格第一行（标题行）。  
2. **当日発送 CSV**：
   * 将空白 **JANコード** 补齐；
   * 删除 JAN 后面的 `-` 及其后内容；
   * 若有运费，请把运费加到单价列；
   * 完成后以 **UTF-8** 编码保存。
3. **手动添加店铺名（复制用）**：
   * 価格.com決済 (クレジットカード)
   * 株式会社テマック
   * 株式会社 GIGA
   * ハーマンズ株式会社
   * 株式会社アウトレットプラザ
   * 千葉入金
    """
)

warehouse = st.sidebar.selectbox("选择仓库", ("通販倉庫", "なんば倉庫"))
orders_file = st.sidebar.file_uploader("上传当日発送 CSV", type=["csv"])
inv_file    = st.sidebar.file_uploader("上传库存余额表 XLSX", type=["xlsx", "xls"])
run_btn = st.sidebar.button("🚀 生成出库单")

OUTPUT_STORES = [
    "販売一丁目 Qoo10店",
    "販売一丁目 Amazon店",
    "販売一丁目 Yahoo！ショッピング店",
    "販売一丁目（楽天）",
    "ニューライフ",
    "販売一丁目 Wowma店",
    "販売一丁目【本店】",     # ← 这行一定要有！括号、空格都一致
    "価格.com決済 (クレジットカード)", # ← 新添加カーゴ　10月8日
    "株式会社テマック", # ← 新添加株式会社テマック　10月14日
    "株式会社 GIGA", # ← 新添加株式会社 GIGA　10月14日
    "ハーマンズ株式会社", # ← 新添加ハーマンズ株式会社　10月14日
    "株式会社アウトレットプラザ", # ← 新添加株式会社アウトレットプラザ　11月11日
    "千葉入金", # ← 新添加千葉入金　2026年3月19日
]

# ----------- Processing Logic -----------
def run_make_outbound(orders_path: str, inv_path: str, warehouse: str, outdir: Path):
    """调用 make_outbound.py 生成出库单文件（输出定向到 outdir）"""
    script = Path(__file__).parent / "make_outbound.py"
    if not script.exists():
        st.error("未找到 make_outbound.py，请将脚本与 app.py 放在同一目录。")
        return
    # 新增 outdir 传参
    cmd = [sys.executable, str(script), orders_path, inv_path, warehouse, str(outdir)]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        st.error(f"脚本执行失败:\n{result.stderr}")
    else:
        st.success("生成完成！")
        st.code(result.stdout)

if run_btn:
    if not orders_file or not inv_file:
        st.warning("请先上传当日発送 CSV 和 库存余额表！")
    else:
        with st.spinner("正在处理，请稍候 …"):
            # 临时目录
            with tempfile.TemporaryDirectory() as tmpdir:
                tmpdir = Path(tmpdir)
                orders_path = tmpdir / orders_file.name
                inv_path    = tmpdir / inv_file.name
                orders_path.write_bytes(orders_file.getvalue())
                inv_path.write_bytes(inv_file.getvalue())

                # 运行脚本，并指定输出目录
                run_make_outbound(str(orders_path), str(inv_path), warehouse, tmpdir)

                # ↓↓↓ 下面这一整段缩进进来 ↓↓↓
                # 先收集所有文件内容
                download_files = []
                store_rows = {}

                for store in OUTPUT_STORES:
                    files = list(tmpdir.glob(f"{store}+*.xlsx"))
                    if files:
                        fpath = files[0]
                        with open(fpath, "rb") as f:
                            file_bytes = f.read()
                        download_files.append((store, fpath.name, file_bytes))
                        # 统计数据行数
                        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
                        ws = wb.active
                        row_count = ws.max_row - 1  # 减去标题行
                        store_rows[store] = row_count
                    else:
                        download_files.append((store, None, None))
                        store_rows[store] = 0

                # 显示每个店铺的数量
                st.subheader("各店铺导出行数")
                for store in OUTPUT_STORES:
                    st.markdown(f"**{store}** ： {store_rows[store]} 行" if store_rows[store] > 0 else f"<span style='color:red'>【{store}】今日无订单</span>", unsafe_allow_html=True)
                
                # 一键下载全部 zip
                if any(fname for _, fname, _ in download_files):
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zipf:
                        for store, fname, file_bytes in download_files:
                            if fname and file_bytes:
                                zipf.writestr(fname, file_bytes)
                    st.download_button(
                        label="📦 下载全部出库单 (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="全部出库单.zip",
                        mime="application/zip",
                        key="download_zip"
                    )
